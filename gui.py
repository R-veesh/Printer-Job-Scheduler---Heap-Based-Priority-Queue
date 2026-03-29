import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import os
import getpass
import math

from printer_scheduler import PrinterScheduler


def count_pages(filepath):
    """Detect page count from a file. Returns (pages, method_description)."""
    ext = os.path.splitext(filepath)[1].lower()

    try:
        if ext == ".pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            pages = len(reader.pages)
            return pages, "PDF pages detected"

        elif ext in (".docx",):
            from docx import Document
            doc = Document(filepath)
            # Estimate: ~250 words per page
            text = " ".join(p.text for p in doc.paragraphs)
            word_count = len(text.split())
            pages = max(1, math.ceil(word_count / 250))
            return pages, f"Estimated from {word_count} words"

        elif ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            pages = len(wb.sheetnames)
            wb.close()
            return pages, f"{pages} sheet(s) detected"

        elif ext == ".csv":
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                line_count = sum(1 for _ in f)
            pages = max(1, math.ceil(line_count / 50))  # ~50 rows per page
            return pages, f"Estimated from {line_count} rows"

        elif ext == ".txt":
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                line_count = sum(1 for _ in f)
            pages = max(1, math.ceil(line_count / 55))  # ~55 lines per page
            return pages, f"Estimated from {line_count} lines"

        elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tiff"):
            return 1, "Image = 1 page"

        else:
            # Fallback: estimate by file size (~3 KB per page)
            size_kb = os.path.getsize(filepath) / 1024
            pages = max(1, math.ceil(size_kb / 3))
            return pages, f"Estimated from file size ({size_kb:.0f} KB)"

    except Exception as e:
        return None, str(e)


class PrinterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Printer Job Scheduler — Heap-Based Priority Queue")
        self.root.geometry("960x700")
        self.root.minsize(860, 620)
        self.root.configure(bg="#f0f2f5")

        self.scheduler = PrinterScheduler()
        self.printing = False
        self.auto_print = False

        self._setup_styles()
        self._build_ui()
        self._refresh_all()

    # ------------------------------------------------------------------ styles
    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("Title.TLabel", font=("Segoe UI", 18, "bold"),
                         background="#2c3e50", foreground="white")
        style.configure("Header.TLabel", font=("Segoe UI", 11, "bold"),
                         background="#f0f2f5")
        style.configure("Status.TLabel", font=("Segoe UI", 10),
                         background="#f0f2f5")

        style.configure("Add.TButton", font=("Segoe UI", 10, "bold"))
        style.configure("Print.TButton", font=("Segoe UI", 10, "bold"))
        style.configure("Cancel.TButton", font=("Segoe UI", 10))

        style.configure("Treeview", font=("Segoe UI", 10), rowheight=26)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))

    # ----------------------------------------------------------- build widgets
    def _build_ui(self):
        # ---- Title bar ----
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=50)
        title_frame.pack(fill="x")
        title_frame.pack_propagate(False)
        ttk.Label(title_frame, text="  🖨  Printer Job Scheduler",
                  style="Title.TLabel").pack(side="left", padx=10, pady=8)

        # ---- Main container ----
        main = tk.Frame(self.root, bg="#f0f2f5")
        main.pack(fill="both", expand=True, padx=12, pady=8)

        # ---- Left panel: input form ----
        left = tk.LabelFrame(main, text=" Submit New Print Job ",
                              font=("Segoe UI", 11, "bold"),
                              bg="#ffffff", padx=14, pady=10)
        left.pack(side="left", fill="y", padx=(0, 8))

        # -- Document Name + Browse --
        tk.Label(left, text="Document Name:", font=("Segoe UI", 10),
                 bg="#ffffff", anchor="w").pack(fill="x", pady=(8, 2))
        doc_row = tk.Frame(left, bg="#ffffff")
        doc_row.pack(fill="x")
        self.doc_entry = tk.Entry(doc_row, font=("Segoe UI", 10), width=16)
        self.doc_entry.pack(side="left", fill="x", expand=True)
        self.browse_btn = tk.Button(doc_row, text="📂 Browse", font=("Segoe UI", 9),
                                     bg="#7f8c8d", fg="white", relief="flat",
                                     cursor="hand2", command=self._browse_file)
        self.browse_btn.pack(side="left", padx=(4, 0), ipady=1)

        self.file_path_var = tk.StringVar(value="")
        self.file_path_label = tk.Label(left, textvariable=self.file_path_var,
                                         font=("Segoe UI", 8), bg="#ffffff",
                                         fg="#7f8c8d", anchor="w", wraplength=220)
        self.file_path_label.pack(fill="x")

        # -- User Name --
        tk.Label(left, text="User Name:", font=("Segoe UI", 10),
                 bg="#ffffff", anchor="w").pack(fill="x", pady=(8, 2))
        self.user_entry = tk.Entry(left, font=("Segoe UI", 10), width=24)
        self.user_entry.pack(fill="x")

        tk.Label(left, text="Priority:", font=("Segoe UI", 10),
                 bg="#ffffff", anchor="w").pack(fill="x", pady=(8, 2))
        self.priority_var = tk.StringVar(value="Medium")
        priority_frame = tk.Frame(left, bg="#ffffff")
        priority_frame.pack(fill="x", pady=2)
        for p in ("High", "Medium", "Low"):
            tk.Radiobutton(priority_frame, text=p, variable=self.priority_var,
                           value=p, font=("Segoe UI", 10), bg="#ffffff",
                           activebackground="#ffffff").pack(side="left", padx=4)

        tk.Label(left, text="Pages:", font=("Segoe UI", 10),
                 bg="#ffffff", anchor="w").pack(fill="x", pady=(8, 2))
        self.pages_entry = tk.Entry(left, font=("Segoe UI", 10), width=24)
        self.pages_entry.pack(fill="x")

        tk.Label(left, text="Print Side:", font=("Segoe UI", 10),
                 bg="#ffffff", anchor="w").pack(fill="x", pady=(8, 2))
        self.side_var = tk.StringVar(value="Single")
        side_frame = tk.Frame(left, bg="#ffffff")
        side_frame.pack(fill="x", pady=2)
        tk.Radiobutton(side_frame, text="Single-Sided", variable=self.side_var,
                       value="Single", font=("Segoe UI", 10), bg="#ffffff",
                       activebackground="#ffffff").pack(side="left", padx=4)
        tk.Radiobutton(side_frame, text="Double-Sided", variable=self.side_var,
                       value="Double", font=("Segoe UI", 10), bg="#ffffff",
                       activebackground="#ffffff").pack(side="left", padx=4)

        btn_frame = tk.Frame(left, bg="#ffffff")
        btn_frame.pack(fill="x", pady=(16, 4))

        self.add_btn = tk.Button(btn_frame, text="➕  Add Job", font=("Segoe UI", 10, "bold"),
                                  bg="#27ae60", fg="white", activebackground="#219a52",
                                  relief="flat", cursor="hand2", command=self._add_job)
        self.add_btn.pack(fill="x", ipady=4)

        self.print_btn = tk.Button(btn_frame, text="🖨  Print Next", font=("Segoe UI", 10, "bold"),
                                    bg="#2980b9", fg="white", activebackground="#2471a3",
                                    relief="flat", cursor="hand2", command=self._print_next)
        self.print_btn.pack(fill="x", ipady=4, pady=(8, 0))

        self.cancel_btn = tk.Button(btn_frame, text="❌  Cancel Selected", font=("Segoe UI", 10),
                                     bg="#c0392b", fg="white", activebackground="#a93226",
                                     relief="flat", cursor="hand2", command=self._cancel_selected)
        self.cancel_btn.pack(fill="x", ipady=4, pady=(8, 0))

        self.auto_btn = tk.Button(btn_frame, text="⏸  Auto Print: OFF", font=("Segoe UI", 10, "bold"),
                                   bg="#7f8c8d", fg="white", activebackground="#6c7a7d",
                                   relief="flat", cursor="hand2", command=self._toggle_auto_print)
        self.auto_btn.pack(fill="x", ipady=4, pady=(8, 0))

        # ---- Right panel: tables ----
        right = tk.Frame(main, bg="#f0f2f5")
        right.pack(side="left", fill="both", expand=True)

        # -- Queue table --
        queue_label_frame = tk.Frame(right, bg="#f0f2f5")
        queue_label_frame.pack(fill="x")
        ttk.Label(queue_label_frame, text="Print Queue (Highest Priority First)",
                  style="Header.TLabel").pack(side="left")
        self.queue_count_label = ttk.Label(queue_label_frame, text="0 jobs",
                                            style="Status.TLabel")
        self.queue_count_label.pack(side="right")

        queue_frame = tk.Frame(right)
        queue_frame.pack(fill="both", expand=True, pady=(4, 8))

        cols = ("ID", "Document", "User", "Priority", "Pages", "Side", "Status")
        self.queue_tree = ttk.Treeview(queue_frame, columns=cols,
                                        show="headings", selectmode="browse")
        for c in cols:
            self.queue_tree.heading(c, text=c)
        self.queue_tree.column("ID", width=40, anchor="center")
        self.queue_tree.column("Document", width=170)
        self.queue_tree.column("User", width=100)
        self.queue_tree.column("Priority", width=70, anchor="center")
        self.queue_tree.column("Pages", width=50, anchor="center")
        self.queue_tree.column("Side", width=60, anchor="center")
        self.queue_tree.column("Status", width=70, anchor="center")

        q_scroll = ttk.Scrollbar(queue_frame, orient="vertical",
                                  command=self.queue_tree.yview)
        self.queue_tree.configure(yscrollcommand=q_scroll.set)
        self.queue_tree.pack(side="left", fill="both", expand=True)
        q_scroll.pack(side="right", fill="y")

        # -- Completed table --
        comp_label_frame = tk.Frame(right, bg="#f0f2f5")
        comp_label_frame.pack(fill="x")
        ttk.Label(comp_label_frame, text="Completed / Cancelled Jobs",
                  style="Header.TLabel").pack(side="left")
        self.comp_count_label = ttk.Label(comp_label_frame, text="0 jobs",
                                           style="Status.TLabel")
        self.comp_count_label.pack(side="right")

        comp_frame = tk.Frame(right)
        comp_frame.pack(fill="both", expand=True, pady=(4, 0))

        self.comp_tree = ttk.Treeview(comp_frame, columns=cols,
                                       show="headings", selectmode="browse")
        for c in cols:
            self.comp_tree.heading(c, text=c)
        self.comp_tree.column("ID", width=40, anchor="center")
        self.comp_tree.column("Document", width=170)
        self.comp_tree.column("User", width=100)
        self.comp_tree.column("Priority", width=70, anchor="center")
        self.comp_tree.column("Pages", width=50, anchor="center")
        self.comp_tree.column("Side", width=60, anchor="center")
        self.comp_tree.column("Status", width=70, anchor="center")

        c_scroll = ttk.Scrollbar(comp_frame, orient="vertical",
                                  command=self.comp_tree.yview)
        self.comp_tree.configure(yscrollcommand=c_scroll.set)
        self.comp_tree.pack(side="left", fill="both", expand=True)
        c_scroll.pack(side="right", fill="y")

        # ---- Bottom status bar ----
        self.status_var = tk.StringVar(value="Ready")
        status_bar = tk.Label(self.root, textvariable=self.status_var,
                               font=("Segoe UI", 9), bg="#2c3e50", fg="#ecf0f1",
                               anchor="w", padx=10)
        status_bar.pack(fill="x", side="bottom")

        # ---- Progress bar (hidden until printing) ----
        self.progress = ttk.Progressbar(self.root, mode="determinate")

    # ----------------------------------------------------------- actions
    def _toggle_auto_print(self):
        self.auto_print = not self.auto_print
        if self.auto_print:
            self.auto_btn.config(text="▶  Auto Print: ON", bg="#8e44ad",
                                 activebackground="#7d3c98")
            self.status_var.set("Auto Print enabled — jobs will print automatically.")
            self._try_auto_print()
        else:
            self.auto_btn.config(text="⏸  Auto Print: OFF", bg="#7f8c8d",
                                 activebackground="#6c7a7d")
            self.status_var.set("Auto Print disabled.")

    def _try_auto_print(self):
        """Start printing the next job if auto-print is on and printer is idle."""
        if self.auto_print and not self.printing and not self.scheduler.queue.is_empty():
            self.printing = True
            self.print_btn.config(state="disabled")
            threading.Thread(target=self._print_worker, daemon=True).start()

    def _browse_file(self):
        filetypes = [
            ("All Files", "*.*"),
            ("PDF Files", "*.pdf"),
            ("Word Documents", "*.doc *.docx"),
            ("Text Files", "*.txt"),
            ("Images", "*.png *.jpg *.jpeg *.bmp"),
            ("Spreadsheets", "*.xls *.xlsx *.csv"),
        ]
        path = filedialog.askopenfilename(
            title="Select a Document to Print",
            filetypes=filetypes,
        )
        if path:
            filename = os.path.basename(path)
            self.doc_entry.delete(0, tk.END)
            self.doc_entry.insert(0, filename)
            self.file_path_var.set(path)

            # Auto-detect page count
            pages, info = count_pages(path)
            if pages is not None:
                self.pages_entry.delete(0, tk.END)
                self.pages_entry.insert(0, str(pages))
                self.status_var.set(f"File: {filename}  |  Pages: {pages} ({info})")
            else:
                self.status_var.set(f"File: {filename}  |  Could not detect pages: {info}")

            # Auto-fill user name from OS if empty
            if not self.user_entry.get().strip():
                self.user_entry.delete(0, tk.END)
                self.user_entry.insert(0, getpass.getuser())

            # Auto-suggest priority based on file type
            ext = os.path.splitext(filename)[1].lower()
            if ext == ".pdf":
                self.priority_var.set("High")
            elif ext in (".docx", ".doc", ".xlsx", ".xls", ".csv"):
                self.priority_var.set("Medium")
            elif ext in (".txt", ".png", ".jpg", ".jpeg", ".bmp", ".gif"):
                self.priority_var.set("Low")

    def _add_job(self):
        doc = self.doc_entry.get().strip()
        user = self.user_entry.get().strip()
        priority = self.priority_var.get()
        pages_str = self.pages_entry.get().strip()

        if not doc:
            messagebox.showwarning("Validation", "Document name cannot be empty.")
            return
        if not user:
            messagebox.showwarning("Validation", "User name cannot be empty.")
            return
        try:
            pages = int(pages_str)
            if pages <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Validation", "Pages must be a positive number.")
            return

        job = self.scheduler.add_job(doc, user, priority, pages)
        job.print_side = self.side_var.get()
        self.status_var.set(f"Job #{job.job_id} \"{job.document_name}\" added  (Priority: {job.priority}, {job.print_side}-Sided)")

        # clear inputs
        self.doc_entry.delete(0, tk.END)
        self.user_entry.delete(0, tk.END)
        self.pages_entry.delete(0, tk.END)
        self.priority_var.set("Medium")
        self.side_var.set("Single")
        self.file_path_var.set("")

        self._refresh_all()
        self._try_auto_print()

    def _print_next(self):
        if self.printing:
            return
        if self.scheduler.queue.is_empty():
            messagebox.showinfo("Queue Empty", "No jobs in the print queue.")
            return
        self.printing = True
        self.print_btn.config(state="disabled")
        threading.Thread(target=self._print_worker, daemon=True).start()

    def _print_worker(self):
        job = self.scheduler.queue.extract_max()
        job.status = "Printing"
        self.scheduler.current_job = job
        total = job.pages

        self.root.after(0, lambda: self.progress.pack(fill="x", side="bottom", before=self.root.winfo_children()[-1]))
        self.root.after(0, lambda: self.progress.configure(maximum=total, value=0))
        self.root.after(0, lambda: self.status_var.set(
            f"Printing Job #{job.job_id} \"{job.document_name}\"  ({getattr(job, 'print_side', 'Single')}-Sided)  \u2014  page 0/{total}"))
        self.root.after(0, self._refresh_all)

        import time
        for i in range(1, total + 1):
            time.sleep(0.25)
            self.root.after(0, lambda v=i: self.progress.configure(value=v))
            self.root.after(0, lambda v=i: self.status_var.set(
                f"Printing Job #{job.job_id} \"{job.document_name}\"  ({getattr(job, 'print_side', 'Single')}-Sided)  \u2014  page {v}/{total}"))

        job.status = "Completed"
        self.scheduler.completed_jobs.append(job)
        self.scheduler.current_job = None

        self.root.after(0, lambda: self.progress.pack_forget())
        self.root.after(0, lambda: self.status_var.set(
            f"Job #{job.job_id} \"{job.document_name}\" printed successfully."))
        self.root.after(0, lambda: self.print_btn.config(state="normal"))
        self.root.after(0, self._refresh_all)
        self.printing = False
        # Auto-print next job if enabled
        self.root.after(500, self._try_auto_print)

    def _cancel_selected(self):
        selected = self.queue_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection", "Select a job from the queue to cancel.")
            return
        item = self.queue_tree.item(selected[0])
        job_id = int(item["values"][0])

        job = self.scheduler.cancel_job(job_id)
        if job is None:
            messagebox.showerror("Error", f"Job #{job_id} not found.")
        else:
            self.status_var.set(f"Job #{job.job_id} \"{job.document_name}\" cancelled.")
        self._refresh_all()

    # ----------------------------------------------------------- refresh
    def _refresh_all(self):
        # Queue table
        for row in self.queue_tree.get_children():
            self.queue_tree.delete(row)
        jobs = self.scheduler.view_queue()
        for j in jobs:
            tag = j.priority.lower()
            self.queue_tree.insert("", "end",
                                    values=(j.job_id, j.document_name, j.user_name,
                                            j.priority, j.pages,
                                            getattr(j, 'print_side', 'Single'),
                                            j.status),
                                    tags=(tag,))
        self.queue_tree.tag_configure("high", background="#fadbd8")
        self.queue_tree.tag_configure("medium", background="#fdebd0")
        self.queue_tree.tag_configure("low", background="#d5f5e3")
        self.queue_count_label.config(text=f"{len(jobs)} job(s)")

        # Completed table
        for row in self.comp_tree.get_children():
            self.comp_tree.delete(row)
        completed = self.scheduler.view_completed()
        for j in completed:
            tag = "done" if j.status == "Completed" else "cancelled"
            self.comp_tree.insert("", "end",
                                   values=(j.job_id, j.document_name, j.user_name,
                                           j.priority, j.pages,
                                           getattr(j, 'print_side', 'Single'),
                                           j.status),
                                   tags=(tag,))
        self.comp_tree.tag_configure("done", background="#d4efdf")
        self.comp_tree.tag_configure("cancelled", background="#f5b7b1")
        self.comp_count_label.config(text=f"{len(completed)} job(s)")


def main():
    root = tk.Tk()
    app = PrinterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
